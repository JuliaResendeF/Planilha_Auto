import openpyxl 
from openpyxl import load_workbook
import datetime 

exel_file = r"C:\Users\PMC\Desktop\exel project\TABELA_python.xlsx"
wb = load_workbook(exel_file)

celula_encontrada = None
conteudo_procurado = int(input("Digite o cod "))
planilha=wb["VIVO"]


for i in planilha.iter_rows():
    for celula in i:
        if celula.value == conteudo_procurado:
            celula_encontrada = celula
            break


def Insert_Values(cod,conteudo_preencher):
    numero_linha = cod
    linha = planilha[numero_linha]
    coluna_vazia = None
    for celula in linha:
        if celula.value is None:
            coluna_vazia = celula.column
            break

    if coluna_vazia:
        planilha.cell(row=numero_linha, column=coluna_vazia, value=conteudo_preencher)
        print("ff")

Cod = celula_encontrada.row
value = input("Digite a data ")

data = datetime.date.today()


if data.month < 10:
    month = '0' + str(data.month)
else:
    month = str(data.month)
    
value = 'OK '+str(data.day)+'/'+ month +'  ' + value
Insert_Values(Cod,value)
value = input("Digite o valor ")
Insert_Values(Cod,value)        
            
wb.save(exel_file)



import openpyxl 
from openpyxl import load_workbook
import datetime 

exel_file = r"Contas.xlsx"
wb = load_workbook(exel_file)
planilha=wb["CPFL"]
celula_encontrada = None
#conteudo_procurado = int(input("Digite o cod "))


def FindLine(conteudo_procurado,celula_encontrada):
    for i in planilha.iter_rows():
        for celula in i:
            if celula.value == conteudo_procurado:
                celula_encontrada = celula
                return celula_encontrada

def Insert_Values(cod,conteudo_preencher):
    numero_linha = cod
    linha = planilha[numero_linha]
    coluna_vazia = None
    for celula in linha:
        if celula.value is None:
            coluna_vazia = celula.column
            break

    if coluna_vazia:
        planilha.cell(row=numero_linha, column=coluna_vazia, value=conteudo_preencher)
        print("ff")

def PresentDate():
    data = datetime.date.today()
    return data

def AutoXLSX(conteudo_procurado, Vencimento_date,valor, tax):
    conteudo_procurado = conteudo_procurado
    celula = FindLine(conteudo_procurado,None)
    Cod =  celula.row

    date = PresentDate()
    value = 'OK '+str(date.day)+'/'+ str(date.month) +'  ' + Vencimento_date
    Insert_Values(Cod,value)
    value =  int(valor) + int(tax)
    Insert_Values(Cod,value)        
    wb.save(exel_file)

def ficha(conteudo_procurado):
    celula = FindLine(conteudo_procurado,None)
    Ficha= 'A'
    conteudo_coluna_desejada = planilha[f'{Ficha}{celula.row}'].value
    
    return conteudo_coluna_desejada
    





